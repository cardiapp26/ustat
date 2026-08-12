"""Value labels typed into Excel cell notes.

SPSS, Stata and SAS carry value labels inside the file and uSTAT already reads
them. An .xlsx has nowhere to put them, so the coding scheme gets typed into a
note on the header cell — and pandas drops notes entirely, which is how a
column of 0s and 1s arrives with nothing to say what they mean.
"""
from __future__ import annotations

import io

import numpy as np
import openpyxl
import pandas as pd
import pytest
from openpyxl.comments import Comment

from routers.upload import _excel_note_metadata, _parse_code_lines

# Reported from a real thyroid dataset, kept verbatim: note the space after
# "3:" and the lower-case seventh label.
THYROID_NOTE = "\n".join([
    "0:Benign",
    "1:Papiller tiroid kanseri",
    "2:Foliküler tiroid kanseri",
    "3: Hurthle hücreli tiroid kanseri",
    "4:Medüller tiroid kanseri",
    "5:Anaplastik",
    "6:Papiller mikrokarsinom",
    "7:papiller karsinom foliküler varyant",
    "8:Diğer",
])


def _book(headers: list[str], rows: list[list], notes: dict[str, str]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    for ref, text in notes.items():
        ws[ref].comment = Comment(text, "Dr")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _meta(content: bytes) -> tuple[dict, pd.DataFrame]:
    df = pd.read_excel(io.BytesIO(content))
    return _excel_note_metadata(content, df), df


@pytest.fixture()
def thyroid() -> bytes:
    rng = np.random.default_rng(0)
    rows = [[int(rng.integers(0, 9)), int(rng.integers(8, 66))] for _ in range(40)]
    return _book(["Histopatoloji", "Tumor boyutu"], rows, {"A1": THYROID_NOTE})


# ── parsing ────────────────────────────────────────────────────────────────────


def test_the_reported_note_becomes_value_labels(thyroid):
    meta, _ = _meta(thyroid)
    labels = meta["Histopatoloji"]["value_labels"]
    assert len(labels) == 9
    assert labels["0"] == "Benign"
    assert labels["3"] == "Hurthle hücreli tiroid kanseri"   # the space after the colon
    assert labels["7"] == "papiller karsinom foliküler varyant"


def test_a_labelled_column_is_treated_as_categorical(thyroid):
    meta, _ = _meta(thyroid)
    assert meta["Histopatoloji"]["measure"] == "nominal"
    assert "Tumor boyutu" not in meta


@pytest.mark.parametrize("line,code,label", [
    ("0:Benign", "0", "Benign"),
    ("3 = Malign", "3", "Malign"),
    ("2) Foliküler", "2", "Foliküler"),
    ("1\tPapiller", "1", "Papiller"),
    ("1 - Papiller", "1", "Papiller"),
    ("-1: Bilinmiyor", "-1", "Bilinmiyor"),
    ("1.5: ara deger", "1.5", "ara deger"),
])
def test_the_separators_people_actually_type(line, code, label):
    assert _parse_code_lines(line) == {code: label}


def test_a_range_is_not_read_as_a_code(client):
    # "1-2 kez" is a label, not the code 1 meaning "2 kez". A bare hyphen only
    # separates when a space follows it.
    assert _parse_code_lines("1-2 kez tekrarlandi") == {}


def test_lines_that_are_not_coding_lines_are_ignored(client):
    # Excel puts the author's name first and people write a sentence above the
    # list; fighting that is not worth it, skipping it is.
    text = "Dr Yilmaz:\nAsagidaki kodlama kullanildi\n0:Yok\n1:Var\n\nSon guncelleme 2024"
    assert _parse_code_lines(text) == {"0": "Yok", "1": "Var"}


# ── knowing when a note is not a coding scheme ─────────────────────────────────


def test_a_plain_description_is_kept_as_the_variable_label(client):
    content = _book(["Boy"], [[170], [165]], {"A1": "Ayakta olculen boy, cm"})
    meta, _ = _meta(content)
    assert meta["Boy"] == {"label": "Ayakta olculen boy, cm"}
    assert "value_labels" not in meta["Boy"]


def test_codes_that_do_not_describe_the_column_do_not_relabel_it(client):
    """A note whose codes have nothing to do with the values in the column.

    Without the coverage check these two lines would attach as value labels
    and, because a labelled column is read as nominal, turn a continuous
    measurement into a categorical one.
    """
    content = _book(["Yas"], [[i] for i in range(30, 80)],
                    {"A1": "0: kayit yok\n1: tahmini yas"})
    meta, _ = _meta(content)
    assert "value_labels" not in meta.get("Yas", {})
    assert "label" in meta["Yas"]


def test_a_single_pair_is_not_a_coding_scheme(client):
    content = _book(["x"], [[1], [2]], {"A1": "1: bir"})
    meta, _ = _meta(content)
    assert "value_labels" not in meta.get("x", {})


# ── matching notes to the right column ─────────────────────────────────────────


def test_notes_are_matched_by_position_across_a_blank_column(client):
    # pandas keeps an empty spreadsheet column as `Unnamed: n`, so positional
    # matching stays aligned; naive name matching would drift by one.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"], ws["C1"] = "grup", "sonuc"
    for i in range(2, 12):
        ws[f"A{i}"], ws[f"C{i}"] = i % 2, i % 3
    ws["C1"].comment = Comment("0:Yok\n1:Var\n2:Belirsiz", "Dr")
    buf = io.BytesIO(); wb.save(buf)
    meta, df = _meta(buf.getvalue())
    assert "sonuc" in meta and "value_labels" in meta["sonuc"]
    assert "grup" not in meta


def test_a_note_on_a_data_cell_is_used_when_the_header_has_none(client):
    content = _book(["g"], [[0], [1], [0], [1]], {"A2": "0:Kadin\n1:Erkek"})
    meta, _ = _meta(content)
    assert meta["g"]["value_labels"] == {"0": "Kadin", "1": "Erkek"}


def test_the_header_note_wins_over_one_further_down(client):
    content = _book(["g"], [[0], [1], [0]],
                    {"A1": "0:Dogru\n1:Dogru2", "A3": "0:Eski\n1:Eski2"})
    meta, _ = _meta(content)
    assert meta["g"]["value_labels"]["0"] == "Dogru"


def test_a_file_with_no_notes_returns_nothing(client):
    meta, _ = _meta(_book(["a", "b"], [[1, 2], [3, 4]], {}))
    assert meta == {}


def test_an_unreadable_workbook_does_not_break_the_upload(client):
    # .xls, encrypted, or anything openpyxl declines: the file still uploads,
    # just without notes.
    assert _excel_note_metadata(b"not a workbook", pd.DataFrame({"a": [1]})) == {}


# ── through the upload endpoint ────────────────────────────────────────────────


def test_uploading_the_file_surfaces_the_labels(client, thyroid):
    r = client.post("/api/upload/", files={
        "file": ("thyroid.xlsx", thyroid,
                 "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200, r.text
    cols = {c["name"]: c for c in r.json()["columns"]}
    assert cols["Histopatoloji"]["kind"] == "categorical"
    assert cols["Histopatoloji"]["value_labels"]["1"] == "Papiller tiroid kanseri"
    # The measurement beside it is untouched.
    assert cols["Tumor boyutu"]["kind"] == "numeric"
    assert not cols["Tumor boyutu"].get("value_labels")
